import hashlib
import io
import json
import os
import re
import urllib.request
import urllib.error
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlsplit, parse_qs

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.abspath(__file__))
TASKS_PATH = os.path.join(ROOT, "..", "tasks.json")
PLANE_CONFIG_PATH = os.path.join(ROOT, "plane_config.json")
# PLANE_HOST and every other Plane setting are per-install: no org's cookie/workspace/project/user
# is baked into the code. They live only in plane_config.json (gitignored, never shipped) and/or
# env vars, and are filled in per-user via the in-app "Connect to Plane" flow (see
# discover_plane_setup below) — nothing here should ever again hardcode ALT Mobility's own ids.
PLANE_HOST = os.environ.get("PLANE_HOST", "https://plane.alt-mobility.com")
PLANE_PRIORITY_MAP = {"P1": "urgent", "P2": "high", "P3": "medium", "P4": "low"}
STATIC_DIR = os.path.join(ROOT, "static")

STATUSES = ["To Do", "In Progress", "In Review", "Pending", "Done", "Cancelled"]
PRIORITIES = ["P1", "P2", "P3", "P4"]
TYPES = ["Task", "Review"]
DATE_FIELDS = ("discussed_from", "discussed_to", "start_date", "due_date", "done_at", "closed_at", "cancelled_at")
ACTIVITY_LIMIT = 200
ACTIVITY_ORDER = {
    "created": 0,
    "status_snapshot": 1,
    "status_changed": 2,
    "completed": 3,
    "reopened": 4,
    "date_changed": 5,
    "attachment_added": 6,
    "attachment_removed": 7,
    "archived": 8,
    "restored": 9,
    "cancelled": 10,
    "cancel_reason_changed": 11,
}


def now_local():
    return datetime.now().astimezone()


def today_local():
    return now_local().date().isoformat()


def iso_now():
    return now_local().isoformat(timespec="seconds")


def clean_date(v):
    if not v:
        return None
    if isinstance(v, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", v):
        return v
    return None


def date_from_any(*values):
    for value in values:
        if not value:
            continue
        if isinstance(value, str) and re.match(r"^\d{4}-\d{2}-\d{2}", value):
            return value[:10]
    return today_local()


def activity_key(a):
    return (
        a.get("type"),
        a.get("date"),
        a.get("from") or a.get("from_status") or a.get("from_value"),
        a.get("to") or a.get("to_status") or a.get("to_value"),
        a.get("field"),
        json.dumps(a.get("value"), sort_keys=True),
        a.get("inferred", False),
        a.get("note") or a.get("reason") or "",
        a.get("at") or "",
    )


def append_activity(task, activity):
    history = task.setdefault("activity_history", [])
    key = activity_key(activity)
    if any(activity_key(existing) == key for existing in history):
        return False
    history.append(activity)
    sort_history(history)
    if len(history) > ACTIVITY_LIMIT:
        del history[: len(history) - ACTIVITY_LIMIT]
    return True


def sort_history(history):
    history.sort(key=lambda a: (a.get("at") or a.get("date") or "", ACTIVITY_ORDER.get(a.get("type"), 99)))


def make_activity(task, activity_type, *, inferred=False, source="system", date_value=None, at=None, **extra):
    at = at or iso_now()
    day = date_value or date_from_any(at)
    activity = {
        "type": activity_type,
        "date": day,
        "at": at,
        "inferred": inferred,
        "source": source,
    }
    activity.update(extra)
    digest = hashlib.sha1(json.dumps(activity_key(activity), sort_keys=True).encode("utf-8")).hexdigest()[:10]
    activity["id"] = f"{task.get('id', 'task')}:{activity_type}:{day}:{digest}"
    return activity


def load_tasks():
    # First run on a fresh clone/install: no tasks.json yet — create an empty one instead of
    # crashing, so a new user (or a new DMG-style install with a fresh data dir) just works.
    if not os.path.exists(TASKS_PATH):
        data = {"tasks": [], "_meta": {}, "settings": {}}
        migrate_data(data)
        save_tasks(data)
        return data
    with open(TASKS_PATH) as f:
        data = json.load(f)
    migrate_data(data)
    return data


def save_tasks(data):
    data.setdefault("_meta", {})
    data["_meta"]["last_regenerated"] = today_local()
    with open(TASKS_PATH, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def migrate_data(data):
    data.setdefault("_meta", {})
    data["_meta"]["schema_version"] = 2
    data["_meta"].setdefault("activity_migration", "Old tasks receive inferred created/status snapshots only; unknown historical transitions are not invented.")
    data.setdefault("settings", {})
    data["settings"].setdefault("daily_update_backlog_limit", 5)
    data.setdefault("tasks", [])
    for task in data["tasks"]:
        normalize_task(task)


def normalize_task(task):
    task.setdefault("id", "")
    task["title"] = (task.get("title") or "").strip() or "Untitled task"
    task["project"] = as_list(task.get("project"))
    task["tags"] = as_list(task.get("tags"))
    task["discussed_with"] = as_list(task.get("discussed_with"))
    task["attachments"] = as_list(task.get("attachments"))
    task["status"] = task.get("status") if task.get("status") in STATUSES else "To Do"
    task["priority"] = task.get("priority") if task.get("priority") in PRIORITIES else "P3"
    task["type"] = task.get("type") if task.get("type") in TYPES else "Task"
    task["cancel_reason"] = (task.get("cancel_reason") or "").strip()
    for field in DATE_FIELDS:
        task[field] = clean_date(task.get(field))
    if not task.get("discussed_from"):
        task["discussed_from"] = date_from_any(task.get("created_at"), task.get("updated_at"))
    if not task.get("discussed_to"):
        task["discussed_to"] = task.get("discussed_from")
    task["month"] = month_from(task.get("discussed_from"))
    task["updated_at"] = clean_date(task.get("updated_at")) or date_from_any(task.get("updated_ts"), task.get("created_at"), task.get("discussed_from"))
    task["created_at"] = task.get("created_at") or f"{task['updated_at']}T00:00:00"
    task["updated_ts"] = task.get("updated_ts") or task.get("created_at")
    task["archived_at"] = task.get("archived_at") or None
    task["plane_issue_id"] = task.get("plane_issue_id") or None
    task["plane_url"] = task.get("plane_url") or None
    if not isinstance(task.get("activity_history"), list):
        task["activity_history"] = []
    if not task["activity_history"]:
        created_date = date_from_any(task.get("created_at"), task.get("discussed_from"), task.get("updated_at"))
        append_activity(task, make_activity(task, "created", inferred=True, source="migration", date_value=created_date, at=task.get("created_at")))
        status_date = date_from_any(task.get("done_at"), task.get("closed_at"), task.get("updated_at"), task.get("discussed_from"))
        append_activity(
            task,
            make_activity(
                task,
                "status_snapshot",
                inferred=True,
                source="migration",
                date_value=status_date,
                at=task.get("updated_ts"),
                to=task["status"],
            ),
        )
        if task["status"] == "Done" and task.get("done_at"):
            append_activity(
                task,
                make_activity(task, "completed", inferred=True, source="migration", date_value=task["done_at"], at=task.get("updated_ts")),
            )
    sort_history(task["activity_history"])


def next_id(tasks):
    nums = []
    for t in tasks:
        m = re.match(r"T(\d+)", t.get("id", ""))
        if m:
            nums.append(int(m.group(1)))
    n = (max(nums) + 1) if nums else 1
    return f"T{n:03d}"


def as_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return [p.strip() for p in str(v).split(",") if p.strip()]


def month_from(iso_date):
    if not iso_date:
        return now_local().strftime("%B")
    try:
        y, m, d = [int(x) for x in iso_date.split("-")]
        return date(y, m, d).strftime("%B")
    except (ValueError, TypeError):
        return now_local().strftime("%B")


CLOSED_STATUSES = {"Done", "Cancelled"}
STALE_DAYS = 7


def is_closed(t):
    return t.get("status") in CLOSED_STATUSES


def is_overdue(t):
    return not is_closed(t) and bool(t.get("due_date")) and t["due_date"] < today_local()


def is_stale(t):
    if is_closed(t):
        return False
    updated = t.get("updated_at") or ""
    cutoff = (now_local() - timedelta(days=STALE_DAYS)).date().isoformat()
    return updated <= cutoff


def filter_tasks_for_export(tasks, q):
    def qval(key):
        v = q.get(key)
        return v[0] if v else ""

    status = qval("status")
    project = qval("project")
    tag = qval("tag")
    ttype = qval("type")
    date_from = qval("dateFrom")
    date_to = qval("dateTo")
    search = qval("search").lower()
    special = qval("filter")  # today | overdue | high | stale | archived | review | "" (all)

    out = []
    for t in tasks:
        archived = bool(t.get("archived_at"))
        if special == "archived":
            if not archived:
                continue
        else:
            if archived:
                continue
        if status and t.get("status") != status:
            continue
        if project and project not in (t.get("project") or []):
            continue
        if tag and tag not in (t.get("tags") or []):
            continue
        if ttype and (t.get("type") or "Task") != ttype:
            continue
        if date_from and (t.get("updated_at") or "") < date_from:
            continue
        if date_to and (t.get("updated_at") or "") > date_to:
            continue
        if special == "overdue" and not is_overdue(t):
            continue
        if special == "high" and (t.get("priority") or "P3") not in ("P1", "P2"):
            continue
        if special == "stale" and not is_stale(t):
            continue
        if special == "review" and (t.get("type") or "Task") != "Review":
            continue
        if special == "today" and (t.get("updated_at") != today_local()):
            continue
        if search:
            hay = " ".join([
                t.get("title", ""), t.get("notes", ""),
                *(t.get("discussed_with") or []), *(t.get("project") or []), *(t.get("tags") or []),
            ]).lower()
            if search not in hay:
                continue
        out.append(t)
    return out


def load_plane_config():
    cfg = {}
    if os.path.exists(PLANE_CONFIG_PATH):
        with open(PLANE_CONFIG_PATH) as f:
            cfg = json.load(f)
    # Env vars win when set — lets someone self-host this (Docker, a shared server, etc.)
    # without ever touching the UI or the config file, per-instance, no code changes.
    for key, env_name in (
        ("cookie", "PLANE_COOKIE"),
        ("workspace", "PLANE_WORKSPACE"),
        ("project_id", "PLANE_PROJECT_ID"),
        ("assignee_id", "PLANE_ASSIGNEE_ID"),
    ):
        val = os.environ.get(env_name)
        if val:
            cfg[key] = val
    return cfg


def save_plane_config(cfg):
    with open(PLANE_CONFIG_PATH, "w") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)
        f.write("\n")


def extract_plane_cookie(raw):
    """The Plane-cookie settings box expects a raw `Cookie:` header value, but people naturally
    paste a whole 'Copy as cURL' command instead (this project's own working examples are all curls).
    If the input looks like a curl invocation, pull the cookie out of it (`-b '...'`, `--cookie '...'`,
    or `-H 'Cookie: ...'`); otherwise assume it's already a plain cookie string and use it as-is."""
    raw = (raw or "").strip()
    if not raw:
        return raw
    patterns = [
        r"(?:-b|--cookie)\s+'([^']*)'",
        r'(?:-b|--cookie)\s+"([^"]*)"',
        r"-H\s+'Cookie:\s*([^']*)'",
        r'-H\s+"Cookie:\s*([^"]*)"',
    ]
    for pat in patterns:
        m = re.search(pat, raw, re.IGNORECASE)
        if m and m.group(1).strip():
            return m.group(1).strip()
    return raw


def plane_state_id(cfg, status):
    state_name = (cfg.get("status_map") or {}).get(status)
    return (cfg.get("states") or {}).get(state_name)


def _plane_request(cfg, method, path, payload=None):
    """Low-level helper for any Plane API call using the stored cookie. Returns (status, data_or_None, raw_text)."""
    cookie = cfg.get("cookie")
    workspace = cfg.get("workspace")
    project_id = cfg.get("project_id")
    url = f"{PLANE_HOST}{path}"
    body = json.dumps(payload).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")
    req.add_header("Cookie", cookie)
    req.add_header("Origin", PLANE_HOST)
    req.add_header("Referer", f"{PLANE_HOST}/{workspace}/projects/{project_id}/issues/")
    req.add_header("User-Agent", "Mozilla/5.0 (task-tracker integration)")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, (json.loads(raw) if raw else None), raw
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return e.code, None, raw
    except urllib.error.URLError as e:
        return None, None, str(e.reason)


PLANE_GROUP_TO_STATUS = {
    "backlog": "To Do",
    "unstarted": "To Do",
    "started": "In Progress",
    "completed": "Done",
    "cancelled": "Cancelled",
}


def discover_plane_setup(cfg):
    """Auto-fills the parts of the Plane config a person shouldn't have to hand-type: their own
    Plane user id (for assignee_ids) and this project's states (for the status_map). Runs
    automatically the moment a cookie + workspace + project_id are saved and no states are known
    yet — this is what lets a new self-hoster connect Plane from the UI alone, no config-file
    editing, and have it stay correct without re-running anything on every future run."""
    me_status, me_data, me_raw = _plane_request(cfg, "GET", "/api/users/me/")
    if me_status != 200 or not me_data:
        return {"error": f"Could not verify Plane session ({me_status}): {(me_raw or '')[:200]}"}

    workspace = cfg.get("workspace")
    project_id = cfg.get("project_id")
    st_status, st_data, st_raw = _plane_request(
        cfg, "GET", f"/api/workspaces/{workspace}/projects/{project_id}/states/"
    )
    if st_status != 200 or not isinstance(st_data, list):
        return {"error": f"Could not fetch project states ({st_status}): {(st_raw or '')[:200]}"}

    states = {s["name"]: s["id"] for s in st_data if s.get("name") and s.get("id")}

    def norm(s):
        return s.lower().replace(" ", "")

    # Never clobber an already hand-tuned mapping — only fill in statuses that aren't mapped yet.
    # A fresh install has none of these, so every status gets a best-effort default; an existing
    # install keeps whatever it already had (even a manually-corrected one), since a workflow with
    # several near-terminal states (e.g. "Dev QA"/"Prod Review"/"Completed" all grouped as
    # "completed" in Plane) is genuinely ambiguous for a heuristic to guess right every time.
    status_map = dict(cfg.get("status_map") or {})
    for local_status in STATUSES:
        if status_map.get(local_status):
            continue
        match = next((s["name"] for s in st_data if norm(s["name"]) == norm(local_status)), None)
        if not match:
            match = next((s["name"] for s in st_data if local_status.lower() in s["name"].lower()
                          or s["name"].lower() in local_status.lower()), None)
        if not match:
            match = next((s["name"] for s in st_data if PLANE_GROUP_TO_STATUS.get(s.get("group")) == local_status), None)
        if match:
            status_map[local_status] = match

    cfg["assignee_id"] = me_data.get("id")
    cfg["assignee_email"] = me_data.get("email")
    cfg["states"] = states
    cfg["status_map"] = status_map
    return {
        "ok": True,
        "assignee_email": me_data.get("email"),
        "state_count": len(states),
        "status_map": status_map,
    }


def plane_meta_comment_html(task):
    """Builds the 'tracker bookkeeping' block (status/priority/type/project/tags/stakeholders/dates/local id)
    that goes into a Plane *comment* — kept out of the description, which should just hold real notes."""
    rows = []
    rows.append(
        f'<p class="editor-paragraph-block"><strong>Tracker status:</strong> {escape_html_py(task.get("status"))} '
        f'&nbsp; <strong>Priority:</strong> {escape_html_py(task.get("priority"))} '
        f'&nbsp; <strong>Type:</strong> {escape_html_py(task.get("type") or "Task")}</p>'
    )
    proj = ", ".join(task.get("project") or [])
    tags = ", ".join(task.get("tags") or [])
    who = ", ".join(task.get("discussed_with") or [])
    if proj:
        rows.append(f'<p class="editor-paragraph-block"><strong>Project:</strong> {escape_html_py(proj)}</p>')
    if tags:
        rows.append(f'<p class="editor-paragraph-block"><strong>Tags:</strong> {escape_html_py(tags)}</p>')
    if who:
        rows.append(f'<p class="editor-paragraph-block"><strong>Stakeholders:</strong> {escape_html_py(who)}</p>')
    dates = []
    if task.get("discussed_from") or task.get("discussed_to"):
        dates.append(f'Discussed: {task.get("discussed_from") or "?"} to {task.get("discussed_to") or "?"}')
    if task.get("start_date"):
        dates.append(f'Start: {task.get("start_date")}')
    if task.get("due_date"):
        dates.append(f'Due: {task.get("due_date")}')
    if task.get("done_at"):
        dates.append(f'Done: {task.get("done_at")}')
    if task.get("closed_at"):
        dates.append(f'Closed: {task.get("closed_at")}')
    if dates:
        rows.append(f'<p class="editor-paragraph-block"><strong>Dates:</strong> {escape_html_py(" | ".join(dates))}</p>')
    if task.get("cancel_reason"):
        rows.append(f'<p class="editor-paragraph-block"><strong>Cancel reason:</strong> {escape_html_py(task.get("cancel_reason"))}</p>')
    rows.append(f'<p class="editor-paragraph-block"><em>Local tracker id: {escape_html_py(task.get("id"))}</em></p>')
    return "".join(rows)


def create_plane_issue(task):
    cfg = load_plane_config()
    cookie = cfg.get("cookie")
    workspace = cfg.get("workspace")
    project_id = cfg.get("project_id")
    if not cookie or not workspace or not project_id:
        return {"error": "Plane is not configured yet — set the cookie in Plane settings first."}
    assignee_id = cfg.get("assignee_id")
    if not assignee_id:
        return {"error": "Plane connected but your user id wasn't detected yet — reopen Plane settings and save again."}

    notes = (task.get("notes") or "").strip()
    desc_html = f'<p class="editor-paragraph-block">{escape_html_py(notes)}</p>' if notes else ""
    payload = {
        "project_id": project_id,
        "type_id": None,
        "name": task.get("title", "Untitled task")[:255],
        "description_html": desc_html,
        "estimate_point": None,
        "state_id": plane_state_id(cfg, task.get("status")),
        "parent_id": None,
        "priority": PLANE_PRIORITY_MAP.get(task.get("priority") or "P3", "none"),
        "assignee_ids": [assignee_id],
        "label_ids": [],
        "cycle_id": None,
        "module_ids": [],
        "start_date": task.get("discussed_from") or None,
        "target_date": task.get("due_date") or None,
    }
    status, data, raw = _plane_request(cfg, "POST", f"/api/workspaces/{workspace}/projects/{project_id}/issues/", payload)
    if status is None:
        return {"error": f"Could not reach Plane: {raw}"}
    if status < 200 or status >= 300 or not data:
        return {"error": f"Plane API error {status}: {raw[:300]}"}

    issue_id = data.get("id")
    if not issue_id:
        return {"error": f"Plane didn't return an issue id: {json.dumps(data)[:300]}"}

    # Bookkeeping (status/priority/type/project/tags/stakeholders/dates) goes into a comment,
    # not the description — keeps the description as just the real notes, per user preference.
    comment_html = plane_meta_comment_html(task)
    _plane_request(cfg, "POST", f"/api/workspaces/{workspace}/projects/{project_id}/issues/{issue_id}/comments/", {"comment_html": comment_html})

    plane_url = f"{PLANE_HOST}/{workspace}/projects/{project_id}/issues/{issue_id}/"
    return {"plane_issue_id": issue_id, "plane_url": plane_url}


def update_plane_issue(task):
    """Pushes the task's CURRENT local state onto its already-linked Plane issue.
    Manual/on-demand only — never called automatically on a field edit. Updates the issue's
    core fields + description (notes), then adds a fresh comment with the current bookkeeping
    block so the change shows up in Plane's Activity feed as a dated update, not a silent overwrite."""
    cfg = load_plane_config()
    cookie = cfg.get("cookie")
    workspace = cfg.get("workspace")
    project_id = cfg.get("project_id")
    issue_id = task.get("plane_issue_id")
    if not cookie or not workspace or not project_id:
        return {"error": "Plane is not configured yet — set the cookie in Plane settings first."}
    if not issue_id:
        return {"error": "This task isn't linked to a Plane issue yet — use Send to Plane first."}
    assignee_id = cfg.get("assignee_id")
    if not assignee_id:
        return {"error": "Plane connected but your user id wasn't detected yet — reopen Plane settings and save again."}

    notes = (task.get("notes") or "").strip()
    desc_html = f'<p class="editor-paragraph-block">{escape_html_py(notes)}</p>' if notes else ""
    payload = {
        "name": task.get("title", "Untitled task")[:255],
        "description_html": desc_html,
        "state_id": plane_state_id(cfg, task.get("status")),
        "priority": PLANE_PRIORITY_MAP.get(task.get("priority") or "P3", "none"),
        "assignee_ids": [assignee_id],
        "start_date": task.get("discussed_from") or None,
        "target_date": task.get("due_date") or None,
    }
    status, data, raw = _plane_request(cfg, "PATCH", f"/api/workspaces/{workspace}/projects/{project_id}/issues/{issue_id}/", payload)
    if status is None:
        return {"error": f"Could not reach Plane: {raw}"}
    if status < 200 or status >= 300:
        return {"error": f"Plane API error {status}: {raw[:300]}"}

    comment_html = plane_meta_comment_html(task)
    _plane_request(cfg, "POST", f"/api/workspaces/{workspace}/projects/{project_id}/issues/{issue_id}/comments/", {"comment_html": comment_html})

    return {"ok": True, "plane_url": task.get("plane_url")}


def escape_html_py(s):
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def build_export_xlsx(tasks):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    FONT = "Arial"
    headers = ["ID", "Type", "Title", "Status", "Priority", "Project(s)", "Tags",
               "Discussed With", "Discussed", "Start", "Due", "Done", "Closed",
               "Notes", "Cancel Reason", "Updated"]

    header_fill = PatternFill("solid", fgColor="4D6D8C")
    header_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E3DDD0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fill = {
        "Done": PatternFill("solid", fgColor="E6F0E9"),
        "In Progress": PatternFill("solid", fgColor="E7EDF3"),
        "In Review": PatternFill("solid", fgColor="EFE7F5"),
        "Pending": PatternFill("solid", fgColor="F6ECD9"),
        "To Do": PatternFill("solid", fgColor="EFE9DC"),
        "Cancelled": PatternFill("solid", fgColor="F5E2E0"),
    }

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    rows_sorted = sorted(tasks, key=lambda t: (t.get("updated_ts") or t.get("updated_at") or ""), reverse=True)
    row = 2
    for t in rows_sorted:
        discussed = t.get("discussed_from") or ""
        if t.get("discussed_to") and t.get("discussed_to") != discussed:
            discussed = f"{discussed} - {t['discussed_to']}"
        values = [
            t.get("id", ""), t.get("type") or "Task", t.get("title", ""), t.get("status", ""),
            t.get("priority") or "P3", ", ".join(t.get("project") or []), ", ".join(t.get("tags") or []),
            ", ".join(t.get("discussed_with") or []), discussed, t.get("start_date") or "",
            t.get("due_date") or "", t.get("done_at") or "", t.get("closed_at") or "",
            t.get("notes", ""), t.get("cancel_reason", ""), t.get("updated_at", ""),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 14)))
            cell.border = border
        status_cell = ws.cell(row=row, column=4)
        status_cell.fill = status_fill.get(t.get("status", ""), PatternFill())
        row += 1

    last_row = max(row - 1, 1)
    widths = {1: 7, 2: 9, 3: 52, 4: 12, 5: 9, 6: 24, 7: 16, 8: 20, 9: 18, 10: 11, 11: 11, 12: 11, 13: 11, 14: 46, 15: 24, 16: 11}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    if last_row >= 1:
        table = Table(displayName="ExportedTasks", ref=f"A1:P{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Exported"
    ws2["B1"] = iso_now()
    ws2["A2"] = "Task count"
    ws2["B2"] = len(tasks)
    for cell in (ws2["A1"], ws2["B1"], ws2["A2"], ws2["B2"]):
        cell.font = Font(name=FONT, size=10)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path, content_type):
        with open(path, "rb") as f:
            body = f.read()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b"{}"
        try:
            return json.loads(raw or b"{}")
        except json.JSONDecodeError:
            return {}

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed = urlsplit(self.path)
        path = parsed.path
        if path == "/" or path == "/index.html":
            return self._send_file(os.path.join(STATIC_DIR, "index.html"), "text/html; charset=utf-8")
        if path == "/style.css":
            return self._send_file(os.path.join(STATIC_DIR, "style.css"), "text/css; charset=utf-8")
        if path == "/app.js":
            return self._send_file(os.path.join(STATIC_DIR, "app.js"), "application/javascript; charset=utf-8")
        if path == "/api/tasks":
            return self._send_json(load_tasks())
        if path == "/api/meta":
            return self._send_json({"statuses": STATUSES, "priorities": PRIORITIES, "types": TYPES})
        if path == "/api/plane-config":
            cfg = load_plane_config()
            return self._send_json({
                "configured": bool(cfg.get("cookie") and cfg.get("states")),
                "workspace": cfg.get("workspace", ""),
                "project_id": cfg.get("project_id", ""),
                "assignee_email": cfg.get("assignee_email", ""),
                "states": sorted((cfg.get("states") or {}).keys()),
                "status_map": cfg.get("status_map", {}),
            })
        if path == "/api/export":
            q = parse_qs(parsed.query)
            data = load_tasks()
            filtered = filter_tasks_for_export(data["tasks"], q)
            xlsx_bytes = build_export_xlsx(filtered)
            fname = f"tasks-export-{today_local()}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            return
        self.send_response(404)
        self.end_headers()

    def do_POST(self):
        if self.path == "/api/plane-config":
            body = self._read_body()
            cfg = load_plane_config()
            if "cookie" in body and body["cookie"].strip():
                cfg["cookie"] = extract_plane_cookie(body["cookie"])
            if "workspace" in body and body["workspace"].strip():
                cfg["workspace"] = body["workspace"].strip()
            if "project_id" in body and body["project_id"].strip():
                cfg["project_id"] = body["project_id"].strip()
            if "status_map" in body and isinstance(body["status_map"], dict):
                cfg["status_map"] = body["status_map"]
            # Auto-detect assignee id + project states the moment we have enough to ask Plane —
            # no hand-editing of the config file, and it stays put across restarts/updates since
            # it's written straight into plane_config.json.
            discovery = None
            if cfg.get("cookie") and cfg.get("workspace") and cfg.get("project_id") and not cfg.get("states"):
                discovery = discover_plane_setup(cfg)
            save_plane_config(cfg)
            resp = {"ok": True, "configured": bool(cfg.get("cookie") and cfg.get("states"))}
            if discovery:
                resp["discovery"] = discovery
            return self._send_json(resp)

        m = re.match(r"^/api/tasks/([^/]+)/plane$", self.path)
        if m:
            task_id = m.group(1)
            data = load_tasks()
            found = None
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            if found.get("plane_issue_id"):
                return self._send_json({
                    "already_exists": True,
                    "plane_issue_id": found["plane_issue_id"],
                    "plane_url": found.get("plane_url"),
                })
            result = create_plane_issue(found)
            if "error" in result:
                return self._send_json(result, status=502)
            found["plane_issue_id"] = result["plane_issue_id"]
            found["plane_url"] = result["plane_url"]
            found["updated_at"] = today_local()
            found["updated_ts"] = iso_now()
            normalize_task(found)
            save_tasks(data)
            return self._send_json(found)

        m = re.match(r"^/api/tasks/([^/]+)/plane-update$", self.path)
        if m:
            task_id = m.group(1)
            data = load_tasks()
            found = None
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            result = update_plane_issue(found)
            if "error" in result:
                return self._send_json(result, status=502)
            return self._send_json(result)

        if self.path == "/api/tasks":
            data = load_tasks()
            body = self._read_body()
            today = today_local()
            now = iso_now()
            d_from = body.get("discussed_from") or today
            d_to = body.get("discussed_to") or d_from
            status = body.get("status") if body.get("status") in STATUSES else "To Do"
            task = {
                "id": next_id(data["tasks"]),
                "title": body.get("title", "").strip() or "Untitled task",
                "project": as_list(body.get("project")),
                "month": month_from(d_from),
                "status": status,
                "priority": body.get("priority") if body.get("priority") in PRIORITIES else "P3",
                "type": body.get("type") if body.get("type") in TYPES else "Task",
                "discussed_with": as_list(body.get("discussed_with")),
                "notes": body.get("notes", ""),
                "tags": as_list(body.get("tags")),
                "attachments": as_list(body.get("attachments")),
                "discussed_from": clean_date(d_from) or today,
                "discussed_to": clean_date(d_to) or clean_date(d_from) or today,
                "start_date": clean_date(body.get("start_date")),
                "due_date": clean_date(body.get("due_date")),
                "done_at": today if status == "Done" else None,
                "closed_at": today if status == "Done" else None,
                "cancelled_at": today if status == "Cancelled" else None,
                "cancel_reason": (body.get("cancel_reason") or "").strip(),
                "updated_at": today,
                "created_at": now,
                "updated_ts": now,
                "archived_at": None,
                "activity_history": [],
            }
            append_activity(task, make_activity(task, "created", at=now))
            if status != "To Do":
                append_activity(task, make_activity(task, "status_changed", at=now, from_status=None, to=status))
            if status == "Cancelled":
                append_activity(task, make_activity(task, "cancelled", at=now, reason=task["cancel_reason"]))
            normalize_task(task)
            data["tasks"].append(task)
            save_tasks(data)
            return self._send_json(task, status=201)
        self.send_response(404)
        self.end_headers()

    def do_PATCH(self):
        m = re.match(r"^/api/tasks/([^/]+)/attachments$", self.path)
        if m:
            task_id = m.group(1)
            data = load_tasks()
            body = self._read_body()
            found = None
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    t.setdefault("attachments", [])
                    label = (body.get("label") or body.get("url") or "").strip()
                    if label:
                        t["attachments"].append(label)
                        append_activity(t, make_activity(t, "attachment_added", label=label))
                    now = iso_now()
                    t["updated_at"] = today_local()
                    t["updated_ts"] = now
                    normalize_task(t)
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            save_tasks(data)
            return self._send_json(found)

        m = re.match(r"^/api/tasks/([^/]+)$", self.path)
        if m:
            task_id = m.group(1)
            data = load_tasks()
            body = self._read_body()
            found = None
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    normalize_task(t)
                    now = iso_now()
                    today = today_local()
                    for key in ("title", "notes"):
                        if key in body:
                            t[key] = body[key]
                    if "cancel_reason" in body:
                        old_reason = t.get("cancel_reason", "")
                        new_reason = (body.get("cancel_reason") or "").strip()
                        if old_reason != new_reason:
                            t["cancel_reason"] = new_reason
                            append_activity(t, make_activity(t, "cancel_reason_changed", at=now, from_value=old_reason, to=new_reason))
                    for key in DATE_FIELDS:
                        if key in body:
                            old_value = t.get(key)
                            new_value = clean_date(body.get(key))
                            if old_value != new_value:
                                t[key] = new_value
                                append_activity(t, make_activity(t, "date_changed", at=now, field=key, from_value=old_value, to=new_value))
                    if "status" in body:
                        old_status = t.get("status")
                        new_status = body["status"] if body["status"] in STATUSES else old_status
                        status_note = (body.get("status_note") or "").strip()
                        if old_status != new_status:
                            t["status"] = new_status
                            status_changed_kwargs = {"from_status": old_status, "to": new_status}
                            if status_note:
                                status_changed_kwargs["note"] = status_note
                            append_activity(t, make_activity(t, "status_changed", at=now, **status_changed_kwargs))
                            if new_status == "Done" and not t.get("done_at"):
                                t["done_at"] = today
                                append_activity(t, make_activity(t, "date_changed", at=now, field="done_at", from_value=None, to=t["done_at"]))
                            if new_status == "Done" and not t.get("closed_at"):
                                t["closed_at"] = today
                            if new_status == "Done":
                                append_activity(t, make_activity(t, "completed", at=now))
                            if new_status == "Cancelled":
                                if not t.get("cancelled_at"):
                                    t["cancelled_at"] = today
                                    append_activity(t, make_activity(t, "date_changed", at=now, field="cancelled_at", from_value=None, to=t["cancelled_at"]))
                                if not t.get("closed_at"):
                                    t["closed_at"] = today
                                append_activity(t, make_activity(t, "cancelled", at=now, reason=t.get("cancel_reason", "")))
                            if old_status == "Done" and new_status != "Done":
                                append_activity(t, make_activity(t, "reopened", at=now, from_status=old_status, to=new_status))
                            if old_status == "Cancelled" and new_status != "Cancelled":
                                t["cancelled_at"] = None
                                append_activity(t, make_activity(t, "reopened", at=now, from_status=old_status, to=new_status))
                    if "discussed_from" in body:
                        t["month"] = month_from(t.get("discussed_from"))
                    if "project" in body:
                        t["project"] = as_list(body["project"])
                    if "tags" in body:
                        t["tags"] = as_list(body["tags"])
                    if "discussed_with" in body:
                        t["discussed_with"] = as_list(body["discussed_with"])
                    if "attachments" in body and isinstance(body["attachments"], list):
                        t["attachments"] = body["attachments"]
                    if "priority" in body and body["priority"] in PRIORITIES:
                        t["priority"] = body["priority"]
                    if "type" in body and body["type"] in TYPES:
                        t["type"] = body["type"]
                    if "archived_at" in body:
                        old_archived_at = t.get("archived_at")
                        t["archived_at"] = clean_date(body.get("archived_at"))
                        if old_archived_at != t.get("archived_at"):
                            append_activity(t, make_activity(t, "archived" if t.get("archived_at") else "restored", at=now))
                    t["updated_at"] = today
                    t["updated_ts"] = now
                    normalize_task(t)
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            save_tasks(data)
            return self._send_json(found)
        self.send_response(404)
        self.end_headers()

    def do_DELETE(self):
        m = re.match(r"^/api/tasks/([^/]+)/attachments/(\d+)$", self.path)
        if m:
            task_id, idx = m.group(1), int(m.group(2))
            data = load_tasks()
            found = None
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    if 0 <= idx < len(t.get("attachments", [])):
                        removed = t["attachments"].pop(idx)
                        append_activity(t, make_activity(t, "attachment_removed", label=removed))
                    now = iso_now()
                    t["updated_at"] = today_local()
                    t["updated_ts"] = now
                    normalize_task(t)
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            save_tasks(data)
            return self._send_json(found)

        m = re.match(r"^/api/tasks/([^/]+)$", self.path)
        if m:
            task_id = m.group(1)
            data = load_tasks()
            found = None
            now = iso_now()
            today = today_local()
            for t in data["tasks"]:
                if t["id"] == task_id:
                    found = t
                    normalize_task(t)
                    if not t.get("archived_at"):
                        t["archived_at"] = today
                        append_activity(t, make_activity(t, "archived", at=now))
                    t["updated_at"] = today
                    t["updated_ts"] = now
                    break
            if not found:
                return self._send_json({"error": "not found"}, status=404)
            save_tasks(data)
            return self._send_json(found)
        self.send_response(404)
        self.end_headers()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8787))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Task tracker running at http://127.0.0.1:{port}")
    server.serve_forever()
